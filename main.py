import openpyxl
import json
from create_dir import Create_dir
from data import Excel_data
from  excel_connaction import Excel_Connection


def excel_file_to_json():
    name_xl = input("Введите название файла: ")
    name_table_xl = input("Введите название таблицы: ")
    directori = input('Введите путь до папки: ')

    cr = Create_dir(directori)
    cr.create_dir()

    dict_excel = {}

    data = Excel_data(name_xl, cr.dir)
    # data.get_value('directori_file')

    ex_conn = Excel_Connection(data.ex_name.ex_name)
    ex_file_conn = ex_conn.conn

    ex_file = ex_file_conn[name_table_xl]

    dict_excel[name_table_xl] = {}

    for len_row in range(2, ex_file.max_row+1):
        c = [] # [cel.value for cel in row for row in ex_file.iter_rows(min_row=len_row, min_col=2, max_row=len_row, max_col=ex_file.max_column)] -> not worked ((
        for row in ex_file.iter_rows(min_row=len_row, min_col=2, max_row=len_row, max_col=ex_file.max_column):
            for cel in row:
                if cel.value is None:
                    continue
                c.append({ex_file.cell(row=1, column=cel.column).value : cel.value})

        dict_excel[(name_table_xl)][ex_file.cell(row=len_row, column=1).value] = c
    #print(json.dumps(dict_excel))

    with open(f"{data.ex_name.directori_file}/data_file.json", "w") as write_file:
         json.dump(dict_excel, write_file, indent=4)


def main():
    excel_file_to_json()


if __name__=='__main__':
    main()